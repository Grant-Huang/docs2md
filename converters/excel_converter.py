"""
Excel 转 Markdown 转换器
支持 .xlsx 格式，.xls 旧格式可选支持
"""
import sys
import os
import io
import re
import json

# 设置标准输出和标准错误的编码为utf-8
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8')


def _send_json(obj):
    """发送 JSON 格式的 SSE 消息"""
    print(f'data: {json.dumps(obj, ensure_ascii=False)}\n\n', flush=True)


def _cell_to_str(cell_value):
    """将单元格值转为字符串"""
    if cell_value is None:
        return ''
    if isinstance(cell_value, float):
        if cell_value == int(cell_value):
            return str(int(cell_value))
        return str(cell_value)
    return str(cell_value).replace('\n', ' ').replace('|', '\\|')


def _rows_to_markdown_table(rows):
    """将行数据转换为 markdown 表格"""
    if not rows:
        return ''
    col_count = max(len(rows[i]) for i in range(len(rows)))
    lines = []
    for row in rows:
        cells = [_cell_to_str(c) for c in row]
        # 补齐列数
        while len(cells) < col_count:
            cells.append('')
        lines.append('| ' + ' | '.join(cells[:col_count]) + ' |')
    if len(lines) >= 2:
        sep = '| ' + ' | '.join(['---'] * col_count) + ' |'
        return lines[0] + '\n' + sep + '\n' + '\n'.join(lines[1:])
    return lines[0] if lines else ''


def _convert_xlsx(excel_path, output_format):
    """使用 openpyxl 转换 xlsx 文件"""
    from openpyxl import load_workbook

    wb = load_workbook(excel_path, read_only=True, data_only=True)
    output_parts = []

    for sheet in wb.worksheets:
        sheet_name = sheet.title
        _send_json({"type": "debug", "content": f"正在处理工作表: {sheet_name}"})

        rows = []
        for row in sheet.iter_rows(values_only=True):
            if row is None:
                continue
            cells = list(row)
            # 跳过全空行
            if not any(c is not None and str(c).strip() != '' for c in cells):
                continue
            rows.append(cells)

        if not rows:
            continue

        if output_format == 'markdown':
            output_parts.append(f'## {sheet_name}')
            output_parts.append('')
            output_parts.append(_rows_to_markdown_table(rows))
            output_parts.append('')
        else:
            output_parts.append(f'=== {sheet_name} ===')
            for row in rows:
                output_parts.append('\t'.join(_cell_to_str(c) for c in row))
            output_parts.append('')

    wb.close()
    return output_parts


def _convert_xls(excel_path, output_format):
    """使用 xlrd 转换 xls 文件（可选）"""
    try:
        import xlrd
    except ImportError:
        return None

    wb = xlrd.open_workbook(excel_path)
    output_parts = []

    for sheet in wb.sheets():
        sheet_name = sheet.name
        _send_json({"type": "debug", "content": f"正在处理工作表: {sheet_name}"})

        rows = []
        for row_idx in range(sheet.nrows):
            row = sheet.row_values(row_idx)
            if any(v != '' for v in row):
                rows.append(row)

        if not rows:
            continue

        if output_format == 'markdown':
            output_parts.append(f'## {sheet_name}')
            output_parts.append('')
            output_parts.append(_rows_to_markdown_table(rows))
            output_parts.append('')
        else:
            output_parts.append(f'=== {sheet_name} ===')
            for row in rows:
                output_parts.append('\t'.join(_cell_to_str(c) for c in row))
            output_parts.append('')

    return output_parts


def convert_excel_to_text(excel_path, output_format='markdown'):
    """将 Excel 文件转换为文本，支持 markdown 和纯文本格式"""
    try:
        sys.stdout.reconfigure(encoding='utf-8')
        sys.stderr.reconfigure(encoding='utf-8')

        if not os.path.exists(excel_path):
            _send_json({"type": "error", "content": f"文件不存在: {excel_path}"})
            return None

        _, ext = os.path.splitext(excel_path)
        ext = ext.lower()

        file_size = os.path.getsize(excel_path)
        _send_json({"type": "debug", "content": f"文件大小: {file_size} 字节"})

        if file_size == 0:
            _send_json({"type": "error", "content": "文件为空"})
            return None

        _send_json({"type": "debug", "content": "正在打开 Excel 文件..."})

        if ext == '.xlsx':
            try:
                output_parts = _convert_xlsx(excel_path, output_format)
            except ImportError:
                _send_json({
                    "type": "error",
                    "content": "缺少 openpyxl 库，请运行: pip install openpyxl"
                })
                return None
        elif ext == '.xls':
            output_parts = _convert_xls(excel_path, output_format)
            if output_parts is None:
                _send_json({
                    "type": "error",
                    "content": (
                        "不支持 .xls 格式，请使用 .xlsx 格式或安装 xlrd: "
                        "pip install xlrd"
                    )
                })
                return None
        else:
            _send_json({
                "type": "error",
                "content": f"不支持的文件格式: {ext}"
            })
            return None

        if not output_parts:
            _send_json({
                "type": "error",
                "content": "警告：没有提取到任何内容"
            })
            return None

        final_text = '\n'.join(output_parts).strip()
        final_text = re.sub(r'\n\s*\n\s*\n', '\n\n', final_text)

        _send_json({
            "type": "debug",
            "content": f"提取内容长度: {len(final_text)} 字符"
        })

        _send_json({"type": "complete", "content": final_text})
        return final_text

    except Exception as e:
        error_msg = f"转换过程中出错: {str(e)}"
        _send_json({"type": "error", "content": error_msg})
        import traceback
        _send_json({
            "type": "error",
            "content": f"错误详情: {traceback.format_exc()}"
        })
        return None


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print(
            "Usage: python excel_converter.py <excel_file_path> [format]",
            file=sys.stderr
        )
        print("format: 'markdown' (default) or 'text'", file=sys.stderr)
        sys.exit(1)

    excel_path = sys.argv[1]
    output_format = sys.argv[2] if len(sys.argv) > 2 else 'markdown'

    if output_format not in ['markdown', 'text']:
        print(
            "Error: format must be either 'markdown' or 'text'",
            file=sys.stderr
        )
        sys.exit(1)

    result = convert_excel_to_text(excel_path, output_format)
    if result:
        sys.exit(0)
    else:
        sys.exit(1)
